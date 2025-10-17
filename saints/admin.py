from django.contrib import admin
from django.utils.html import format_html
from django.forms.models import BaseInlineFormSet
from django.core.exceptions import ValidationError
from .models import Image
from .models import (
    Saint,
    HagiographyOrWrittenSource,
    AspectsOfCult,
    ArchaeologicalAndArchitecturalEvidence,
    Image,
    Bibliography,
)
import openpyxl
from django.http import HttpResponse


import csv
from django.http import HttpResponse
from openpyxl import Workbook


from django.http import HttpResponse
from openpyxl import Workbook


def export_saints_full(modeladmin, request, queryset):
    wb = Workbook()

    # --- Sheet 1: Saints ---
    ws = wb.active
    ws.title = "Saints"
    ws.append(
        [
            "ID",
            "Name",
            "Gender",
            "Type of Saint",
            "Act Period 1",
            "Act Period 2",
            "Act Period Text",
            "Region of Birth",
            "Birth Lat",
            "Birth Lng",
            "Region of Burial",
            "Burial Lat",
            "Burial Lng",
            "Feast Day",
        ]
    )
    for saint in queryset:
        ws.append(
            [
                saint.id,
                saint.name,
                saint.gender,
                saint.type_of_saint,
                saint.act_period_1,
                saint.act_period_2,
                saint.act_period_text,
                saint.region_of_birth,
                saint.region_of_birth_latitude,
                saint.region_of_birth_longitude,
                saint.region_of_burial,
                saint.region_of_burial_latitude,
                saint.region_of_burial_longitude,
                saint.feast_day,
            ]
        )

    # --- Sheet 2: Images ---
    ws = wb.create_sheet("Images")
    ws.append(
        [
            "Saint ID",
            "Saint Name",
            "Image File",
            "Is Default",
            "Material",
            "Site/Monument",
            "Latitude",
            "Longitude",
            "Date",
            "Century",
            "Location of Portrait",
            "Inscription",
            "Description",
        ]
    )
    for saint in queryset:
        for img in saint.images.all():
            ws.append(
                [
                    saint.id,
                    saint.name,
                    img.image.url if img.image else "No file",
                    img.is_default,
                    img.material_technique,
                    img.site_monument,
                    img.site_monument_latitude,
                    img.site_monument_longitude,
                    img.date,
                    img.century,
                    img.location_of_portrait,
                    img.inscription_text,
                    img.description,
                ]
            )

    # --- Sheet 3: Hagiographies ---
    ws = wb.create_sheet("Hagiographies")
    ws.append(
        [
            "Saint ID",
            "Saint Name",
            "Literary Type",
            "Date/Century",
            "Title",
            "Liturgical Type",
            "Liturgical Date",
            "Liturgical Title",
        ]
    )
    for saint in queryset:
        for h in saint.hagiographies.all():
            ws.append(
                [
                    saint.id,
                    saint.name,
                    h.literary_source_type,
                    h.date_or_century_of_source,
                    h.title_literary_text,
                    h.liturgical_texts_source_type,
                    h.date_or_century_of_liturgical_source,
                    h.title_liturgical_text,
                ]
            )

    # --- Sheet 4: Cult Aspects ---
    ws = wb.create_sheet("Cult Aspects")
    ws.append(
        [
            "Saint ID",
            "Saint Name",
            "Cult Place",
            "Latitude",
            "Longitude",
            "Pilgrimage Site",
            "Status",
            "Activities",
            "Relics/Objects",
            "Miracles",
        ]
    )
    for saint in queryset:
        for c in saint.cult_aspects.all():
            ws.append(
                [
                    saint.id,
                    saint.name,
                    c.cult_place,
                    c.cult_place_latitude,
                    c.cult_place_longitude,
                    c.pilgrimage_site,
                    c.status,
                    c.activities_accompanying_cult,
                    c.relics_cult_related_objects,
                    c.miracles,
                ]
            )

    # --- Sheet 5: Archaeology ---
    ws = wb.create_sheet("Archaeology")
    ws.append(
        [
            "Saint ID",
            "Saint Name",
            "Building Type",
            "Latitude",
            "Longitude",
            "Location",
            "Date 1",
            "Date 2",
            "Date Text",
            "Condition",
        ]
    )
    for saint in queryset:
        for a in saint.archaeological_evidence.all():
            ws.append(
                [
                    saint.id,
                    saint.name,
                    a.cult_buildings,
                    a.cult_buildings_latitude,
                    a.cult_buildings_longitude,
                    a.location,
                    a.date_1,
                    a.date_2,
                    a.date_text,
                    a.current_condition,
                ]
            )

    # --- Sheet 6: Bibliography ---
    ws = wb.create_sheet("Bibliography")
    ws.append(["Saint ID", "Saint Name", "Reference"])
    for saint in queryset:
        for b in saint.bibliographies.all():
            ws.append([saint.id, saint.name, b.reference])

    # --- Sheet 7: Buildings ---
    ws = wb.create_sheet("Buildings")
    ws.append(
        [
            "Saint ID",
            "Saint Name",
            "Building Name",
            "Type",
            "Location",
            "Construction Date",
            "Description",
            "Is Active",
        ]
    )
    for saint in queryset:
        for b in saint.buildings.all():
            ws.append(
                [
                    saint.id,
                    saint.name,
                    b.name,
                    b.building_type,
                    b.location,
                    b.construction_date,
                    b.description,
                    b.is_active,
                ]
            )

    # --- Return response ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="saints_full_export.xlsx"'
    wb.save(response)
    return response


export_saints_full.short_description = "Export full saints dataset (with relationships)"


class ImageInlineFormSet(BaseInlineFormSet):
    def clean(self):
        super().clean()
        defaults = 0
        for form in self.forms:
            if form.cleaned_data.get("DELETE"):
                continue
            if form.cleaned_data.get("is_default"):
                defaults += 1
        if defaults > 1:
            raise ValidationError(
                "Only one image can be marked as default for a saint."
            )


class ImageInline(admin.TabularInline):
    model = Image
    extra = 1
    fields = (
        "image",
        "is_default",
        "material_technique",
        "site_monument",
        "site_monument_latitude",
        "site_monument_longitude",
        "date",
        "century",
        "location_of_portrait",
        "location_of_portrait_text",
        "iconographic_type",
        "has_inscription",
        "inscription_text",
        "description",
    )


class HagiographyInline(admin.TabularInline):
    model = HagiographyOrWrittenSource
    extra = 1


class CultInline(admin.TabularInline):
    model = AspectsOfCult
    extra = 1


class ArchaeologyInline(admin.TabularInline):
    model = ArchaeologicalAndArchitecturalEvidence
    extra = 1


class BibliographyInline(admin.TabularInline):
    model = Bibliography
    extra = 1


# ===== Main admin =====
@admin.register(Saint)
class SaintAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "name",
        "gender",
        "type_of_saint",
        "act_period_1",
        "act_period_2",
        "image_preview",
    )
    list_filter = ("gender", "type_of_saint", "act_period_1", "act_period_2")
    search_fields = ("id", "name", "region_of_birth", "region_of_burial")
    ordering = ("name",)
    list_display_links = ("id", "name")
    readonly_fields = ("id",)
    actions = [export_saints_full]

    inlines = [
        ImageInline,
        HagiographyInline,
        CultInline,
        ArchaeologyInline,
        BibliographyInline,
    ]

    def image_preview(self, obj):
        if not obj:
            return "No image"

        # Prefer the default image, fallback to the first
        default_image = obj.images.filter(is_default=True).first()
        image_obj = default_image or obj.images.first()

        if image_obj and image_obj.image:
            return format_html(
                '<img src="{}" style="max-height: 50px; max-width: 50px; object-fit: cover;" />',
                image_obj.image.url,
            )
        return "No image"

    image_preview.short_description = "Image Preview"


# ===== Other admins =====
@admin.register(HagiographyOrWrittenSource)
class HagiographyOrWrittenSourceAdmin(admin.ModelAdmin):
    list_display = (
        "title_literary_text",
        "saint",
        "literary_source_type",
        "date_or_century_of_source",
    )
    list_filter = ("literary_source_type",)
    search_fields = ("title_literary_text", "saint__name")

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "saint":
            kwargs["queryset"] = Saint.objects.all().order_by("name")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(AspectsOfCult)
class AspectsOfCultAdmin(admin.ModelAdmin):
    list_display = ("saint", "cult_place", "pilgrimage_site", "status")
    list_filter = ("pilgrimage_site", "status")
    search_fields = ("saint__name", "cult_place")

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "saint":
            kwargs["queryset"] = Saint.objects.all().order_by("name")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(ArchaeologicalAndArchitecturalEvidence)
class ArchaeologicalAndArchitecturalEvidenceAdmin(admin.ModelAdmin):
    list_display = (
        "saint",
        "cult_buildings",
        "date_1",
        "date_2",
        "date_text",
        "location",
    )
    list_filter = ("cult_buildings",)
    search_fields = ("saint__name", "location")

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "saint":
            kwargs["queryset"] = Saint.objects.all().order_by("name")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(Image)
class ImageAdmin(admin.ModelAdmin):
    list_display = (
        "saint",
        "image",
        "material_technique",
        "site_monument",
        "date",
        "century",
        "location_of_portrait",
        "location_description",
        "iconographic_type",
        "has_inscription",
        "inscription_text",
        "description",
    )
    list_filter = ("material_technique", "century", "iconographic_type")
    search_fields = ("saint__name", "site_monument")

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "saint":
            kwargs["queryset"] = Saint.objects.all().order_by("name")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(Bibliography)
class BibliographyAdmin(admin.ModelAdmin):
    list_display = ("saint", "reference")
    search_fields = ("saint__name", "reference")

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "saint":
            kwargs["queryset"] = Saint.objects.all().order_by("name")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
